import pandas as pd
from django.db.models import Case, When, CharField, Value
import logging

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO


from django.http import HttpResponse

from abc import ABC
from django.db.models import Q
from django.db import transaction
from commons.middlewares.exception import BadRequestException

from staffing_status.models import StaffingStatus
from personnel_database.models.subsatker import SubSatKer
from personnel_database.models.pangkat import Pangkat

CUSTOM_ORDER = [
    "PIMPINAN",
    "DIT KAMSEL",
    "DIT GAKKUM",
    "DIT REGIDENT",
    "BAG OPS",
    "BAG RENMIN",
    "BAG TIK",
    "SIKEU",
    "TAUD"
]
ORDER_DICT = {name: i for i, name in enumerate(CUSTOM_ORDER)}

class StaffingService(ABC):
    
    @classmethod
    @transaction.atomic
    def update_staffing_status(cls, **kwargs) :
        logger = logging.getLogger('general')
        subsatker = kwargs.pop("satker")
        data = kwargs.pop("data")
        
        logger.info(f"Updating staffing status for satker: {subsatker}")
        
        # Get the SubSatKer object first
        try:
            subsatker_obj = SubSatKer.objects.get(nama=subsatker)
        except SubSatKer.DoesNotExist:
            error_msg = f"Satker '{subsatker}' not found in SubSatKer. Available satkers: {list(SubSatKer.objects.values_list('nama', flat=True))}"
            logger.error(error_msg)
            raise BadRequestException(error_msg)

        # Log available pangkat
        available_pangkat = Pangkat.objects.values_list('nama', flat=True)
        logger.info(f"Available pangkat in database: {list(available_pangkat)}")
        
        # Get or create staffing status for each pangkat
        for i in data:
            i = dict(i)
            pangkat_name = i['pangkat']
            logger.info(f"Processing pangkat: {pangkat_name}")
            
            # Find or create pangkat object
            pangkat_obj, pangkat_created = Pangkat.objects.get_or_create(
                nama=pangkat_name,
                defaults={'tipe': 'POLRI' if pangkat_name not in ['IV', 'III', 'II/I'] else 'PNS POLRI'}
            )
            
            if pangkat_created:
                logger.info(f"Created new pangkat: {pangkat_name}")
            
            # Get or create staffing status
            staffing_status, created = StaffingStatus.objects.get_or_create(
                nama=pangkat_name,
                subsatker=subsatker_obj,
                defaults={'dsp': i['dsp']}
            )
            
            if created:
                staffing_status.pangkat.add(pangkat_obj)
                logger.info(f"Created new staffing status for {subsatker} - {pangkat_name}")
            else:
                staffing_status.dsp = i['dsp']
                staffing_status.save()
                logger.info(f"Updated existing staffing status for {subsatker} - {pangkat_name}")

        return cls.get_staffing_status()

    @classmethod
    def update_data(cls, data: StaffingStatus, dsp: int) :
        data.dsp = dsp
        data.save()

    @classmethod
    def get_staffing_status(cls) :
        data = []

        satker_list = sorted(SubSatKer.objects.all(), key=lambda x: ORDER_DICT.get(x.nama, len(CUSTOM_ORDER)))
        for i in satker_list :
            temp_polri = cls.get_staffing_data("POLRI", i)
            temp_pns_polri = cls.get_staffing_data("PNS POLRI", i)
            temp = cls.data_wrapper(i.nama, temp_polri, temp_pns_polri)
            data.append(temp)

        return data

    @classmethod
    def get_staffing_data(cls, tipe: str, subsatker:SubSatKer) :
        data = {
                tipe: {
                    "jumlah" : {
                        "dsp" : 0,
                        "rill" : 0
                    }
                }}
        
        staffing_status_list = StaffingStatus.objects.filter(Q(pangkat__tipe = tipe) & Q(subsatker=subsatker)).distinct()
        for i in staffing_status_list :
            message = ""
            if(i.dsp > i.rill) :
                message = f"Subsatker {i.subsatker} dengan Pangkat {i.nama} kekurangan {i.dsp - i.rill} personil"
            elif(i.dsp < i.rill) :
                message = f"Subsatker {i.subsatker} dengan Pangkat {i.nama} kelebihan {i.rill - i.dsp} personil"
  
            temp = {
                "dsp" : i.dsp,
                "rill" : i.rill,
                "message" : message
            }
            data[tipe][i.nama] = temp
            data[tipe]['jumlah']['dsp'] = data[tipe]['jumlah']['dsp'] + i.dsp
            data[tipe]['jumlah']['rill'] = data[tipe]['jumlah']['rill'] + i.rill

        return data
    
    @classmethod
    def data_wrapper(cls, satker, polri: dict, pns_polri: dict) :
        data = {"satker" : satker,
                "keterangan" : {
                    "dsp" : polri["POLRI"]["jumlah"]["dsp"] + pns_polri["PNS POLRI"]["jumlah"]["dsp"],
                    "rill" : polri["POLRI"]["jumlah"]["rill"] + pns_polri["PNS POLRI"]["jumlah"]["rill"]
                },
                **polri, **pns_polri
        }
        return data
    
    @classmethod
    def get_total_by_pangkat(cls) :
        data = {
            "POLRI" : {
                "dsp": 0,
                "rill": 0
            },
            "PNS POLRI" : {
                "dsp" : 0,
                "rill" : 0
            },
            "Keterangan" : {
                "dsp" : 0,
                "rill" : 0,
            }
        }

        staffing_status_list = StaffingStatus.objects.all().distinct().annotate(
                tipe=Case(
                When(pangkat__tipe='PNS POLRI', then=Value("PNS POLRI")),
                default=Value("POLRI"),
                output_field=CharField(),
                )
        )

        for i in staffing_status_list :
            if(not data.get(i.nama, None)) :
                data[i.nama] = {
                    'dsp' : 0,
                    'rill' : 0,
                }

            data[i.nama]['dsp'] = data[i.nama]['dsp'] +  i.dsp
            data[i.nama]['rill'] = data[i.nama]['rill'] + i.rill

            data[i.tipe]['dsp'] = data[i.tipe]['dsp'] + i.dsp
            data[i.tipe]['rill'] = data[i.tipe]['rill'] + i.rill

            data['Keterangan']['dsp'] = data['Keterangan']['dsp'] + i.dsp
            data['Keterangan']['rill'] = data['Keterangan']['rill'] + i.rill

        return data
    
    @classmethod
    def export_csv_file(cls):
        # Define the MultiIndex for columns
        columns = pd.MultiIndex.from_tuples([
            ('', 'No'), ('', 'SatKer'),
            ('IRJEN', 'DSP'), ('IRJEN', 'RIIL'),
            ('BRIGJEN', 'DSP'), ('BRIGJEN', 'RIIL'),
            ('KOMBES', 'DSP'), ('KOMBES', 'RIIL'),
            ('AKBP', 'DSP'), ('AKBP', 'RIIL'),
            ('KOMPOL', 'DSP'), ('KOMPOL', 'RIIL'),
            ('AKP', 'DSP'), ('AKP', 'RIIL'),
            ('IP', 'DSP'), ('IP', 'RIIL'),
            ('BRIGADIR', 'DSP'), ('BRIGADIR', 'RIIL'),
            ('Jumlah', 'DSP'), ('Jumlah', 'RIIL'),
            ('IV', 'DSP'), ('IV', 'RIIL'),
            ('III', 'DSP'), ('III', 'RIIL'),
            ('II/I', 'DSP'), ('II/I', 'RIIL'),
            ('Jumlah PNS', 'DSP'), ('Jumlah PNS', 'RIIL'), 
            ('Ket', 'DSP'), ('Ket', 'RIIL')
        ], names=['', ''])
         
        # List kategori pangkat
        polri_pangkat = ['IRJEN', 'BRIGJEN', 'KOMBES', 'AKBP', 'KOMPOL', 'AKP', 'IP', 'BRIGADIR']
        pns_polri_pangkat = ['IV', 'III', 'II/I']
        
        # Ambil semua SubSatKer dalam urutan custom yang ditentukan
        satker_list = sorted(SubSatKer.objects.all(), key=lambda x: ORDER_DICT.get(x.nama, len(CUSTOM_ORDER)))
        
        # Dapatkan data staffing status dari database
        data = cls.get_staffing_status()

        # Create the DataFrame
        df = pd.DataFrame(columns=columns)

        # Isi dataframe dengan data
        for idx, satker_data in enumerate(data, start=1):
            satker = satker_data['satker']
            row = [idx, satker]
            
            # Proses data POLRI
            polri_dsp_total = 0
            polri_riil_total = 0
            
            for level in polri_pangkat:
                dsp = 0
                riil = 0
                if 'POLRI' in satker_data and level in satker_data['POLRI']:
                    dsp = satker_data['POLRI'][level]['dsp']
                    riil = satker_data['POLRI'][level]['rill']
                
                polri_dsp_total += dsp
                polri_riil_total += riil
                row.extend([dsp, riil])
            
            # Tambahkan total POLRI
            if 'POLRI' in satker_data and 'jumlah' in satker_data['POLRI']:
                polri_dsp_total = satker_data['POLRI']['jumlah']['dsp']
                polri_riil_total = satker_data['POLRI']['jumlah']['rill']
            
            row.extend([polri_dsp_total, polri_riil_total])
            
            # Proses data PNS POLRI
            pns_dsp_total = 0
            pns_riil_total = 0
            
            for level in pns_polri_pangkat:
                dsp = 0
                riil = 0
                if 'PNS POLRI' in satker_data and level in satker_data['PNS POLRI']:
                    dsp = satker_data['PNS POLRI'][level]['dsp']
                    riil = satker_data['PNS POLRI'][level]['rill']
                
                pns_dsp_total += dsp
                pns_riil_total += riil
                row.extend([dsp, riil])
            
            # Tambahkan total PNS POLRI
            if 'PNS POLRI' in satker_data and 'jumlah' in satker_data['PNS POLRI']:
                pns_dsp_total = satker_data['PNS POLRI']['jumlah']['dsp']
                pns_riil_total = satker_data['PNS POLRI']['jumlah']['rill']
                
            row.extend([pns_dsp_total, pns_riil_total])
            
            # Tambahkan kolom keterangan
            if 'keterangan' in satker_data:
                ket_dsp = satker_data['keterangan']['dsp']
                ket_riil = satker_data['keterangan']['rill']
                row.extend([ket_dsp, ket_riil])
            else:
                row.extend(['...', '...'])
            
            # Tambahkan row ke dataframe
            df.loc[len(df)] = row

        # Calculate totals
        totals = ['Jumlah', '']
        for i in range(2, len(df.columns)):
            totals.append(df.iloc[:, i].sum())
        
        # Add totals row to DataFrame
        df.loc[len(df)] = totals

        # Create an Excel workbook and worksheet
        wb = Workbook()
        ws = wb.active
        
        # Tambahkan header utama
        ws.merge_cells('C1:T1')
        ws.merge_cells('U1:AB1')
        ws['C1'].value = 'POLRI'
        ws['U1'].value = 'PNS POLRI'
        ws['C1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['U1'].alignment = Alignment(horizontal='center', vertical='center')

        # Append the dataframe to the worksheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Merge cells for the header
        ws.merge_cells('C2:D2')  # IRJEN
        ws.merge_cells('E2:F2')  # BRIGJEN
        ws.merge_cells('G2:H2')  # KOMBES
        ws.merge_cells('I2:J2')  # AKBP
        ws.merge_cells('K2:L2')  # KOMPOL
        ws.merge_cells('M2:N2')  # AKP
        ws.merge_cells('O2:P2')  # IP
        ws.merge_cells('Q2:R2')  # BRIGADIR
        ws.merge_cells('S2:T2')  # Jumlah POLRI
        ws.merge_cells('U2:V2')  # IV
        ws.merge_cells('W2:X2')  # III
        ws.merge_cells('Y2:Z2')  # II/I
        ws.merge_cells('AA2:AB2')  # Jumlah PNS
        ws.merge_cells('AC2:AD2')  # Keterangan
        
        # Set header values
        headers = ['IRJEN', 'BRIGJEN', 'KOMBES', 'AKBP', 'KOMPOL', 'AKP', 'IP', 'BRIGADIR', 'Jumlah', 'IV', 'III', 'II/I', 'Jumlah', 'Ket']
        for i, header in enumerate(headers, start=3):
            cell = ws.cell(row=2, column=i*2-3)
            cell.value = header
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Align the subheaders (DSP/RIIL)
        for col_num in range(3, len(columns) + 1):
            cell = ws.cell(row=3, column=col_num)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width

        # Save the workbook to a virtual file
        virtual_workbook = BytesIO()
        wb.save(virtual_workbook)

        # Create HTTP response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="staffing-status.xlsx"'
        response.write(virtual_workbook.getvalue())

        return response